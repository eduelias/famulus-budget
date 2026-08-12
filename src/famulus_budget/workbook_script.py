import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import os
BD = os.environ.get('BUDGET_DIR', '/budget')
rows = json.load(open(os.path.join(BD, 'pnl_rows.json')))

INCOME = ['Salary (Schuberg Philis)','Child benefit (SVB)','Tax refunds','Other income']
EXPENSE = ['Mortgage','Energy & water','Taxes (municipal & national)','Insurance','Car & transport',
 'Groceries','Household & drugstore','Childcare (GO)','Kids lessons (ballet & piano)','Health & medical',
 'Telecom & internet','Subscriptions & digital','Entertainment & eating out','Clothing & personal',
 'Online shopping','Transfers & Tikkie','Work (reimbursed)','Services & other']
ONEOFF = ['Renovation depot (one-off)','Renovation (one-off)']
MONTHS = [f"2026-{m:02d}" for m in range(1,13)]
MLABEL = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

wb = Workbook()

# ---------------- Transactions ----------------
ws_t = wb.create_sheet('Transactions')
hdr = ['Date','Month','Source','Merchant','Category','Amount']
ws_t.append(hdr)
for r in rows:
    ws_t.append(r)
n = len(rows)+1
tab = Table(displayName='Txns', ref=f"A1:F{n}")
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws_t.add_table(tab)
for c,w in zip('ABCDEF',[12,10,8,38,28,12]): ws_t.column_dimensions[c].width = w
for row in ws_t.iter_rows(min_row=2, min_col=6, max_col=6):
    row[0].number_format = '#,##0.00'

# ---------------- PnL ----------------
ws = wb.active; ws.title = 'PnL'
BOLD = Font(bold=True); WHITE_B = Font(bold=True, color='FFFFFF')
HDRFILL = PatternFill('solid', fgColor='1F4E78')
SECFILL = PatternFill('solid', fgColor='DDEBF7')
NETFILL = PatternFill('solid', fgColor='FFF2CC')
EUR = '#,##0;[Red]-#,##0'

ws['A1'] = 'P&L 2026 — Saleh household (bank + credit card, cash basis)'; ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Jan–Jul filled from statements. Add new rows in Transactions and months update automatically. Credit-card settlements excluded (card purchases itemized instead).'
ws['A2'].font = Font(italic=True, size=9)

r0 = 4
ws.cell(r0, 1, 'Category').font = WHITE_B; ws.cell(r0,1).fill = HDRFILL
for i, ml in enumerate(MLABEL):
    c = ws.cell(r0, 2+i, ml); c.font = WHITE_B; c.fill = HDRFILL; c.alignment = Alignment(horizontal='center')
c = ws.cell(r0, 14, 'Total'); c.font = WHITE_B; c.fill = HDRFILL; c.alignment = Alignment(horizontal='center')

def put_block(start, title, cats, sign):
    r = start
    ws.cell(r, 1, title).font = BOLD; ws.cell(r,1).fill = SECFILL
    for i in range(13): ws.cell(r, 2+i).fill = SECFILL
    r += 1
    first = r
    for cat in cats:
        ws.cell(r, 1, cat)
        for i, mon in enumerate(MONTHS):
            col = get_column_letter(2+i)
            f = f'=SUMIFS(Txns[Amount],Txns[Month],"{mon}",Txns[Category],"{cat}")'
            ws.cell(r, 2+i, f'={sign}SUMIFS(Txns[Amount],Txns[Month],"{mon}",Txns[Category],"{cat}")' if sign else f)
            ws.cell(r, 2+i).number_format = EUR
        ws.cell(r, 14, f'=SUM(B{r}:M{r})').number_format = EUR
        r += 1
    # subtotal
    ws.cell(r, 1, f'Total {title.lower()}').font = BOLD
    for i in range(12):
        col = get_column_letter(2+i)
        ws.cell(r, 2+i, f'=SUM({col}{first}:{col}{r-1})').number_format = EUR
        ws.cell(r, 2+i).font = BOLD
    ws.cell(r, 14, f'=SUM(B{r}:M{r})').number_format = EUR; ws.cell(r,14).font = BOLD
    return r

inc_end = put_block(r0+1, 'Income', INCOME, '')
exp_end = put_block(inc_end+2, 'Expenses', EXPENSE, '-')

net_r = exp_end + 2
ws.cell(net_r, 1, 'NET (regular)').font = Font(bold=True, size=12)
for i in range(12):
    col = get_column_letter(2+i)
    ws.cell(net_r, 2+i, f'={col}{inc_end}-{col}{exp_end}').number_format = EUR
    ws.cell(net_r, 2+i).font = BOLD; ws.cell(net_r, 2+i).fill = NETFILL
ws.cell(net_r, 14, f'=SUM(B{net_r}:M{net_r})').number_format = EUR
ws.cell(net_r, 14).font = BOLD; ws.cell(net_r, 14).fill = NETFILL
ws.cell(net_r, 1).fill = NETFILL

oo_r = net_r + 2
ws.cell(oo_r, 1, 'One-offs (renovation/solar, excluded from NET above)').font = Font(italic=True, bold=True)
r = oo_r + 1
for cat in ONEOFF:
    ws.cell(r, 1, cat).font = Font(italic=True)
    for i, mon in enumerate(MONTHS):
        ws.cell(r, 2+i, f'=SUMIFS(Txns[Amount],Txns[Month],"{mon}",Txns[Category],"{cat}")').number_format = EUR
        ws.cell(r, 2+i).font = Font(italic=True)
    ws.cell(r, 14, f'=SUM(B{r}:M{r})').number_format = EUR
    r += 1

ws.column_dimensions['A'].width = 34
for i in range(12): ws.column_dimensions[get_column_letter(2+i)].width = 9
ws.column_dimensions['N'].width = 11
ws.freeze_panes = 'B5'

# ---------------- Budget ----------------
wsb = wb.create_sheet('Budget')
wsb['A1'] = 'Budget follow-up 2026'; wsb['A1'].font = Font(bold=True, size=14)
wsb['A2'] = 'Change the budget amounts as you like. Actuals pull from the Transactions sheet. Green = within budget, red = over.'
wsb['A2'].font = Font(italic=True, size=9)

budgets = [('Groceries', 1000), ('Entertainment & eating out', 150), ('Household & drugstore', 250), ('Subscriptions & digital', 100), ('Childcare (GO)', 154), ('Kids lessons (ballet & piano)', 85)]
r = 4
for cat, bud in budgets:
    wsb.cell(r, 1, cat).font = Font(bold=True, size=12); wsb.cell(r,1).fill = SECFILL
    for i in range(13): wsb.cell(r, 2+i).fill = SECFILL
    r += 1
    wsb.cell(r, 1, 'Month')
    for i, ml in enumerate(MLABEL): wsb.cell(r, 2+i, ml).font = BOLD
    wsb.cell(r, 14, 'Total').font = BOLD
    r += 1
    wsb.cell(r, 1, 'Budget')
    for i in range(12):
        wsb.cell(r, 2+i, bud).number_format = EUR
    wsb.cell(r, 14, f'=SUM(B{r}:M{r})').number_format = EUR
    bud_row = r
    r += 1
    wsb.cell(r, 1, 'Actual')
    for i, mon in enumerate(MONTHS):
        wsb.cell(r, 2+i, f'=-SUMIFS(Txns[Amount],Txns[Month],"{mon}",Txns[Category],"{cat}")').number_format = EUR
    wsb.cell(r, 14, f'=SUM(B{r}:M{r})').number_format = EUR
    act_row = r
    r += 1
    wsb.cell(r, 1, 'Left / (over)')
    for i in range(12):
        col = get_column_letter(2+i)
        wsb.cell(r, 2+i, f'={col}{bud_row}-{col}{act_row}').number_format = EUR
    wsb.cell(r, 14, f'=N{bud_row}-N{act_row}').number_format = EUR
    rng = f'B{r}:N{r}'
    wsb.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', fgColor='F8CBAD')))
    wsb.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=PatternFill('solid', fgColor='C6EFCE')))
    r += 2

wsb.column_dimensions['A'].width = 16
for i in range(12): wsb.column_dimensions[get_column_letter(2+i)].width = 9
wsb.column_dimensions['N'].width = 11

# order sheets
wb.move_sheet('Budget', offset=-1)
wb.move_sheet('PnL', offset=-2)

out = os.path.join(BD, 'PnL-Budget-2026.xlsx')
wb.save(out)
print('saved', out)
